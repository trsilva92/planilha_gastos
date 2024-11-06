import os
import yaml
from openpyxl import load_workbook
from src.convert_pdf import main as convert_pdf
from src.convert_xls import main as convert_xls

def main():
    planilha_gastos, tabelas, faturas = carregar_arquivos()

    # Dicionário para armazenar totais por mês
    totais_por_mes = {}

    for file_xslx in faturas:
        mes = file_xslx.split("-")[1].replace('.xlsx', '')
        fatura = load_workbook(f"../resources/faturas/{file_xslx}")
        sheet = fatura.active
        despesas = carregar_despesas(sheet, tabelas)

        # Acumula despesas nas categorias
        if mes not in totais_por_mes:
            totais_por_mes[mes] = {}

        for categoria, valor in despesas.items():
            if categoria == 'OUTROS':
                for descricao, val in valor.items():
                    totais_por_mes[mes].setdefault('OUTROS', {}).setdefault(descricao, 0)
                    totais_por_mes[mes]['OUTROS'][descricao] += val
            else:
                totais_por_mes[mes][categoria] = totais_por_mes[mes].get(categoria, 0) + valor

    # Preenche as abas correspondentes após processar todas as faturas
    for mes, totais in totais_por_mes.items():
        aba_atual = planilha_gastos[mes]  # Referência à aba do mês atual
        preencher_dados(aba_atual, tabelas, totais)

    # Salva a planilha final
    planilha_gastos.save("../resources/Planilha de Gastos 2024.xlsx")

def carregar_arquivos():
    faturas = os.listdir("../resources/faturas")

    if not faturas:
        convert_xls()
        convert_pdf()

    # Carrega a planilha de gastos e as tabelas do YAML
    planilha_gastos = load_workbook("../resources/Planilha de Gastos 2024.xlsx")
    with open('../resources/tabelas.yml', 'r', encoding='utf-8') as file:
        tabelas = yaml.safe_load(file)

    return planilha_gastos, tabelas, faturas

def carregar_despesas(sheet, tabelas):
    totais = {}
    for linha in range(2, sheet.max_row + 1):
        descricao = sheet.cell(row=linha, column=5).value.upper().strip()
        valor = sheet.cell(row=linha, column=9).value

        # Ignora descrições definidas no objeto ignorar em tabelas_template.yml
        if any(ignorar in descricao for ignorar in tabelas["ignorar"]):
            continue

        # Organiza as despesas nas categorias correspondentes
        for nome, dados in tabelas["despesas_variaveis"].items():
            if any(term in descricao for term in dados['termos']):
                totais[nome] = totais.get(nome, 0) + valor
                break
        else:
            totais.setdefault('OUTROS', {})[descricao] = totais.get('OUTROS', {}).get(descricao, 0) + valor

    return totais

def preencher_dados(nova_aba, tabelas, totais):
    # Adiciona entradas
    for celula, info in tabelas["entradas"].items():
        nova_aba[celula] = info['nome']
        nova_aba[f'C{celula[1:]}'] = info['valor']

    # Adiciona despesas fixas
    for celula, info in tabelas["despesas_fixas"].items():
        nova_aba[celula] = info['nome']
        nova_aba[f'F{celula[1:]}'] = info['valor']

    # Adiciona despesas variáveis
    for nome, dados in tabelas["despesas_variaveis"].items():
        col_e = dados['col_e']
        col_f = dados['col_f']

        if nome in totais and isinstance(col_e, str):
            nova_aba[col_e] = nome.replace("_", " ")
            nova_aba[col_f] = totais[nome]

    # Adiciona outras despesas
    linha_inicial = 5
    for descricao, valor in totais.get('OUTROS', {}).items():
        nova_aba[f'I{linha_inicial}'] = descricao
        nova_aba[f'J{linha_inicial}'] = valor
        linha_inicial += 1

if __name__ == "__main__":
    main()
